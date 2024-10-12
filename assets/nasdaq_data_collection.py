import yfinance as yf
import pandas as pd
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment, numbers

start_date = datetime.datetime(1971, 2, 5)
end_date = datetime.datetime(2024, 10, 11)

nasdaq_data = yf.download("^IXIC", start=start_date, end=end_date)

nasdaq_data["변동비"] = nasdaq_data["Adj Close"].pct_change()

nasdaq_data_filtered = nasdaq_data[["Adj Close", "변동비"]].copy()
nasdaq_data_filtered.columns = ["나스닥 지수", "변동비"]

nasdaq_data_filtered.index = nasdaq_data_filtered.index.date

nasdaq_data_filtered["나스닥 지수(근사)"] = (
    nasdaq_data_filtered["나스닥 지수"].round(0).astype(int)
)

nasdaq_data_filtered["일일 수익률(%)"] = nasdaq_data_filtered["변동비"] * 100
nasdaq_data_filtered["일일 수익률(%)"] = nasdaq_data_filtered["일일 수익률(%)"].apply(
    lambda x: f"+{x:.2f}%" if x > 0 else f"{x:.2f}%"
)

nasdaq_data_filtered.index.name = "날짜"
excel_file = "nasdaq_daily_returns.xlsx"
nasdaq_data_filtered.to_excel(excel_file, index=True)

wb = load_workbook(excel_file)
ws = wb.active

for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    adjusted_width = max_length + 2
    ws.column_dimensions[column].width = adjusted_width

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
    for cell in row:
        cell.number_format = "YYYY-MM-DD"
        cell.alignment = Alignment(horizontal="center")

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=4, max_col=5):
    for cell in row:
        cell.alignment = Alignment(horizontal="center")

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=5, max_col=5):
    for cell in row:
        cell.number_format = numbers.FORMAT_PERCENTAGE_00

wb.save(excel_file)
